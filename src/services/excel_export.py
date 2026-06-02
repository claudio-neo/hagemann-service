"""
Exportación Excel — Hagemann Stundenkonto
Genera reporte mensual de saldos de horas en formato .xlsx
"""
import io
from calendar import monthrange
from typing import Optional
from uuid import UUID
from datetime import date, datetime, timedelta

from sqlalchemy.orm import Session
from sqlalchemy import func

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

from ..models.empleado import Empleado, CentroCoste
from ..models.saldo_horas import SaldoHorasMensual
from ..models.fichaje import Fichaje, SegmentoTiempo
from ..models.vacaciones import SolicitudVacaciones, TipoAusencia, EstadoSolicitud, Festivo

MONTH_NAMES = [
    "", "Januar", "Februar", "März", "April", "Mai", "Juni",
    "Juli", "August", "September", "Oktober", "November", "Dezember",
]


# ─── Colores ────────────────────────────────────────────────────────────────
HEADER_BG   = "2C3E50"   # azul oscuro
HEADER_FG   = "FFFFFF"   # blanco
ROW_NEG_BG  = "FDECEA"   # rojo claro → diferencia negativa
TOTAL_BG    = "D5D8DC"   # gris
ALT_ROW_BG  = "EBF5FB"   # azul muy claro (filas pares)

HEADERS = [
    "Personalnr.", "Nachname", "Vorname",
    "Planstunden", "Iststunden", "Differenz",
    "Saldo", "Krankheitst.", "Urlaubst.", "Feiertage",
]


def _make_fill(hex_color: str) -> "PatternFill":
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _thin_border() -> "Border":
    thin = Side(style="thin", color="BDBDBD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _count_ausencia(
    db: Session,
    empleado_id: UUID,
    year: int,
    month: int,
    tipo: str,
) -> int:
    """Cuenta días de ausencia de un tipo dado en el mes."""
    mes_inicio = date(year, month, 1)
    # Último día del mes
    if month == 12:
        mes_fin = date(year + 1, 1, 1)
    else:
        mes_fin = date(year, month + 1, 1)

    result = (
        db.query(func.sum(SolicitudVacaciones.dias))
        .filter(
            SolicitudVacaciones.empleado_id == empleado_id,
            SolicitudVacaciones.estado == EstadoSolicitud.APROBADA,
            SolicitudVacaciones.tipo_ausencia == tipo,
            SolicitudVacaciones.fecha_inicio >= mes_inicio,
            SolicitudVacaciones.fecha_inicio < mes_fin,
        )
        .scalar()
    )
    return int(result or 0)


def _count_festivos(db: Session, year: int, month: int) -> int:
    """Cuenta festivos activos en el mes (días laborables)."""
    mes_inicio = date(year, month, 1)
    if month == 12:
        mes_fin = date(year + 1, 1, 1)
    else:
        mes_fin = date(year, month + 1, 1)

    count = (
        db.query(func.count(Festivo.id))
        .filter(
            Festivo.activo == True,
            Festivo.fecha >= mes_inicio,
            Festivo.fecha < mes_fin,
        )
        .scalar()
    )
    return int(count or 0)


def _fill_sheet(
    ws,
    db: Session,
    year: int,
    month: int,
    empleado_ids: Optional[list] = None,
):
    """Fills a worksheet with the monthly report data."""
    header_fill = _make_fill(HEADER_BG)
    total_fill = _make_fill(TOTAL_BG)
    neg_fill = _make_fill(ROW_NEG_BG)
    alt_fill = _make_fill(ALT_ROW_BG)
    border = _thin_border()

    query = (
        db.query(SaldoHorasMensual, Empleado)
        .join(Empleado, SaldoHorasMensual.empleado_id == Empleado.id)
        .filter(
            SaldoHorasMensual.anio == year,
            SaldoHorasMensual.mes == month,
        )
    )
    if empleado_ids:
        query = query.filter(SaldoHorasMensual.empleado_id.in_(empleado_ids))

    rows = query.order_by(Empleado.id_nummer).all()
    festivos_mes = _count_festivos(db, year, month)

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color=HEADER_FG, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 20

    totals = {k: 0.0 for k in
              ["plan", "ist", "diff", "saldo", "krank", "urlaub", "feiertage"]}

    for row_idx, (saldo, emp) in enumerate(rows, start=2):
        plan = float(saldo.horas_planificadas or 0)
        ist = float(saldo.horas_reales or 0)
        diff = ist - plan
        saldo_val = float(saldo.saldo_final or 0)

        krank = _count_ausencia(db, emp.id, year, month, TipoAusencia.BAJA_MEDICA)
        urlaub = _count_ausencia(db, emp.id, year, month, TipoAusencia.VACACIONES)

        data = [
            emp.id_nummer, emp.apellido or "", emp.nombre,
            round(plan, 2), round(ist, 2), round(diff, 2),
            round(saldo_val, 2), krank, urlaub, festivos_mes,
        ]

        is_negative = diff < 0
        row_fill = neg_fill if is_negative else (alt_fill if row_idx % 2 == 0 else None)

        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="right" if col_idx > 3 else "left",
                vertical="center",
            )
            cell.font = Font(size=10)
            if row_fill:
                cell.fill = row_fill

        totals["plan"] += plan
        totals["ist"] += ist
        totals["diff"] += diff
        totals["saldo"] += saldo_val
        totals["krank"] += krank
        totals["urlaub"] += urlaub
        totals["feiertage"] += festivos_mes

    total_row = len(rows) + 2
    total_data = [
        "GESAMT", "", "",
        round(totals["plan"], 2), round(totals["ist"], 2),
        round(totals["diff"], 2), round(totals["saldo"], 2),
        int(totals["krank"]), int(totals["urlaub"]), int(totals["feiertage"]),
    ]
    for col_idx, value in enumerate(total_data, start=1):
        cell = ws.cell(row=total_row, column=col_idx, value=value)
        cell.fill = total_fill
        cell.font = Font(bold=True, size=10)
        cell.border = border
        cell.alignment = Alignment(
            horizontal="right" if col_idx > 3 else "left",
            vertical="center",
        )
    ws.row_dimensions[total_row].height = 18

    for col_idx in range(1, len(HEADERS) + 1):
        max_len = len(HEADERS[col_idx - 1])
        for r in range(2, total_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    ws.freeze_panes = "A2"


def generar_reporte_mensual(
    db: Session,
    year: int,
    month: int,
    empleado_ids: Optional[list] = None,
) -> bytes:
    """Genera el reporte mensual de saldos de horas en .xlsx (una hoja)."""
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl no está instalado. Añadir a requirements.txt")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Stunden {year}-{month:02d}"
    _fill_sheet(ws, db, year, month, empleado_ids)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def preview_reporte_mensual(
    db: Session,
    year: int,
    month: int,
    empleado_ids: Optional[list] = None,
) -> dict:
    """
    Devuelve preview JSON de los datos que se exportarían en Excel.
    """
    query = (
        db.query(SaldoHorasMensual, Empleado)
        .join(Empleado, SaldoHorasMensual.empleado_id == Empleado.id)
        .filter(
            SaldoHorasMensual.anio == year,
            SaldoHorasMensual.mes == month,
        )
    )
    if empleado_ids:
        query = query.filter(SaldoHorasMensual.empleado_id.in_(empleado_ids))

    rows = query.order_by(Empleado.id_nummer).all()
    festivos_mes = _count_festivos(db, year, month)

    data = []
    for saldo, emp in rows:
        plan = float(saldo.horas_planificadas or 0)
        ist  = float(saldo.horas_reales or 0)
        data.append({
            "personalnr":   emp.id_nummer,
            "nachname":     emp.apellido or "",
            "vorname":      emp.nombre,
            "planstunden":  round(plan, 2),
            "iststunden":   round(ist, 2),
            "differenz":    round(ist - plan, 2),
            "saldo":        round(float(saldo.saldo_final or 0), 2),
            "krankheitstage": _count_ausencia(db, emp.id, year, month, TipoAusencia.BAJA_MEDICA),
            "urlaubstage":    _count_ausencia(db, emp.id, year, month, TipoAusencia.VACACIONES),
            "feiertage":    festivos_mes,
        })

    totals = {
        "planstunden":    round(sum(r["planstunden"] for r in data), 2),
        "iststunden":     round(sum(r["iststunden"] for r in data), 2),
        "differenz":      round(sum(r["differenz"] for r in data), 2),
        "saldo":          round(sum(r["saldo"] for r in data), 2),
        "krankheitstage": sum(r["krankheitstage"] for r in data),
        "urlaubstage":    sum(r["urlaubstage"] for r in data),
    }

    return {
        "year": year,
        "month": month,
        "total_empleados": len(data),
        "festivos_mes": festivos_mes,
        "empleados": data,
        "totales": totals,
    }


# ═══════════════════════════════════════════════════════════════════════════
#  REPORTE POR RANGO DE FECHAS (von – bis)
# ═══════════════════════════════════════════════════════════════════════════

SUMMARY_HEADERS = [
    "Personalnr.", "Nachname", "Vorname",
    "Planstunden", "Iststunden", "Differenz",
    "Saldo (kumuliert)",
    "Krankheitst.", "Urlaubst.", "Feiertage",
]


def _saldo_kumuliert(db: Session, empleado_id, year: int, month: int) -> float:
    """Saldo acumulado (saldo_final, con Übertrag) del mes indicado. 0 si no existe."""
    row = db.query(SaldoHorasMensual.saldo_final).filter(
        SaldoHorasMensual.empleado_id == empleado_id,
        SaldoHorasMensual.anio == year,
        SaldoHorasMensual.mes == month,
    ).first()
    return round(float(row[0]), 2) if row and row[0] is not None else 0.0

DETAIL_HEADERS = [
    "Personalnr.", "Name", "Datum", "Beginn", "Ende",
    "Pause (Min)", "Stunden", "Kostenstelle",
]


def _get_horas_reales_rango(db: Session, empleado_id, fecha_von: date, fecha_bis: date) -> float:
    """Suma horas de fichajes CERRADOS cuya fecha_entrada cae en [von, bis]."""
    inicio = datetime.combine(fecha_von, datetime.min.time())
    fin = datetime.combine(fecha_bis + timedelta(days=1), datetime.min.time())
    result = db.query(
        func.coalesce(func.sum(Fichaje.minutos_trabajados), 0)
    ).filter(
        Fichaje.empleado_id == empleado_id,
        Fichaje.fecha_salida.isnot(None),
        Fichaje.fecha_entrada >= inicio,
        Fichaje.fecha_entrada < fin,
    ).scalar()
    return round(float(result or 0) / 60.0, 2)


def _count_ausencia_rango(db: Session, empleado_id, fecha_von: date, fecha_bis: date, tipo: str) -> int:
    """Días de ausencia aprobada de un tipo cuya fecha_inicio cae en [von, bis]."""
    result = db.query(
        func.sum(SolicitudVacaciones.dias)
    ).filter(
        SolicitudVacaciones.empleado_id == empleado_id,
        SolicitudVacaciones.estado == EstadoSolicitud.APROBADA,
        SolicitudVacaciones.tipo_ausencia == tipo,
        SolicitudVacaciones.fecha_inicio >= fecha_von,
        SolicitudVacaciones.fecha_inicio <= fecha_bis,
    ).scalar()
    return int(result or 0)


def _count_festivos_rango(db: Session, fecha_von: date, fecha_bis: date) -> int:
    """Festivos activos con fecha en [von, bis]."""
    count = db.query(func.count(Festivo.id)).filter(
        Festivo.activo == True,
        Festivo.fecha >= fecha_von,
        Festivo.fecha <= fecha_bis,
    ).scalar()
    return int(count or 0)


def _planstunden_rango(emp: Empleado, fecha_von: date, fecha_bis: date) -> float:
    """
    Prorratea monthly_hours sobre el rango, mes a mes, igual que el cálculo mensual:
    por cada mes tocado → monthly_hours * (días del rango en el mes / días del mes).
    Respeta beginn_berechnung/fecha_alta y fecha_baja del empleado.
    """
    inicio_calculo = emp.beginn_berechnung or emp.fecha_alta
    von, bis = fecha_von, fecha_bis
    if inicio_calculo and inicio_calculo > von:
        von = inicio_calculo
    if emp.fecha_baja and emp.fecha_baja < bis:
        bis = emp.fecha_baja
    if von > bis:
        return 0.0

    monthly = float(emp.monthly_hours or 0)
    total = 0.0
    y, m = von.year, von.month
    while (y, m) <= (bis.year, bis.month):
        dim = monthrange(y, m)[1]
        seg_start = max(von, date(y, m, 1))
        seg_end = min(bis, date(y, m, dim))
        dias = (seg_end - seg_start).days + 1
        if dias > 0:
            total += monthly * (dias / dim)
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return round(total, 2)


def _resumen_empleados_rango(db: Session, fecha_von: date, fecha_bis: date, empleado_ids=None):
    """Construye las filas de resumen por empleado para el rango."""
    q = db.query(Empleado)
    if empleado_ids:
        q = q.filter(Empleado.id.in_(empleado_ids))
    else:
        q = q.filter(Empleado.activo == True)
    empleados = q.order_by(Empleado.id_nummer).all()

    festivos = _count_festivos_rango(db, fecha_von, fecha_bis)
    saldo_y, saldo_m = fecha_bis.year, fecha_bis.month
    rows = []
    for emp in empleados:
        ist = _get_horas_reales_rango(db, emp.id, fecha_von, fecha_bis)
        plan = _planstunden_rango(emp, fecha_von, fecha_bis)
        rows.append({
            "personalnr": emp.id_nummer,
            "nachname": emp.apellido or "",
            "vorname": emp.nombre,
            "planstunden": plan,
            "iststunden": ist,
            "differenz": round(ist - plan, 2),
            "saldo_kumuliert": _saldo_kumuliert(db, emp.id, saldo_y, saldo_m),
            "krankheitstage": _count_ausencia_rango(db, emp.id, fecha_von, fecha_bis, TipoAusencia.BAJA_MEDICA),
            "urlaubstage": _count_ausencia_rango(db, emp.id, fecha_von, fecha_bis, TipoAusencia.VACACIONES),
            "feiertage": festivos,
        })
    return rows, festivos


def _detalle_fichajes_rango(db: Session, fecha_von: date, fecha_bis: date, empleado_ids=None):
    """Una fila por fichaje cerrado en el rango, con centro(s) de coste."""
    inicio = datetime.combine(fecha_von, datetime.min.time())
    fin = datetime.combine(fecha_bis + timedelta(days=1), datetime.min.time())
    q = (
        db.query(Fichaje, Empleado)
        .join(Empleado, Fichaje.empleado_id == Empleado.id)
        .filter(
            Fichaje.fecha_salida.isnot(None),
            Fichaje.fecha_entrada >= inicio,
            Fichaje.fecha_entrada < fin,
        )
    )
    if empleado_ids:
        q = q.filter(Fichaje.empleado_id.in_(empleado_ids))
    fichajes = q.order_by(Empleado.id_nummer, Fichaje.fecha_entrada).all()

    rows = []
    for fich, emp in fichajes:
        centros = sorted({
            seg.centro_coste.nombre
            for seg in fich.segmentos
            if seg.centro_coste is not None
        })
        rows.append({
            "personalnr": emp.id_nummer,
            "name": f"{emp.apellido or ''} {emp.nombre}".strip(),
            "datum": fich.fecha_entrada,
            "beginn": fich.fecha_entrada,
            "ende": fich.fecha_salida,
            "pause": int(fich.minutos_descanso or 0),
            "stunden": round(float(fich.minutos_trabajados or 0) / 60.0, 2),
            "kostenstelle": ", ".join(centros),
        })
    return rows


def _write_header(ws, headers, fill, fg, border):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = fill
        cell.font = Font(bold=True, color=fg, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 20


def _autowidth(ws, headers, last_row):
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])
        for r in range(2, last_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3


def generar_reporte_rango(
    db: Session,
    fecha_von: date,
    fecha_bis: date,
    empleado_ids: Optional[list] = None,
) -> bytes:
    """
    Genera reporte por rango de fechas con 2 hojas:
      1. Zusammenfassung — resumen por empleado (horas reales + ausencias en el rango)
      2. Details — listado de fichajes individuales del rango
    """
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl no está instalado. Añadir a requirements.txt")

    header_fill = _make_fill(HEADER_BG)
    total_fill = _make_fill(TOTAL_BG)
    neg_fill = _make_fill(ROW_NEG_BG)
    alt_fill = _make_fill(ALT_ROW_BG)
    border = _thin_border()

    wb = Workbook()

    # ── Hoja 1: Zusammenfassung ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Zusammenfassung"
    _write_header(ws1, SUMMARY_HEADERS, header_fill, HEADER_FG, border)

    resumen, _ = _resumen_empleados_rango(db, fecha_von, fecha_bis, empleado_ids)
    totals = {k: 0.0 for k in ["plan", "ist", "diff", "saldo", "krank", "urlaub", "feiertage"]}

    for row_idx, r in enumerate(resumen, start=2):
        data = [
            r["personalnr"], r["nachname"], r["vorname"],
            r["planstunden"], r["iststunden"], r["differenz"],
            r["saldo_kumuliert"],
            r["krankheitstage"], r["urlaubstage"], r["feiertage"],
        ]
        row_fill = neg_fill if r["saldo_kumuliert"] < 0 else (alt_fill if row_idx % 2 == 0 else None)
        for col_idx, value in enumerate(data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if col_idx > 3 else "left", vertical="center")
            cell.font = Font(size=10)
            if row_fill:
                cell.fill = row_fill
        totals["plan"] += r["planstunden"]
        totals["ist"] += r["iststunden"]
        totals["diff"] += r["differenz"]
        totals["saldo"] += r["saldo_kumuliert"]
        totals["krank"] += r["krankheitstage"]
        totals["urlaub"] += r["urlaubstage"]
        totals["feiertage"] = r["feiertage"]

    total_row = len(resumen) + 2
    total_data = [
        "GESAMT", "", "",
        round(totals["plan"], 2), round(totals["ist"], 2), round(totals["diff"], 2),
        round(totals["saldo"], 2),
        int(totals["krank"]), int(totals["urlaub"]), int(totals["feiertage"]),
    ]
    for col_idx, value in enumerate(total_data, start=1):
        cell = ws1.cell(row=total_row, column=col_idx, value=value)
        cell.fill = total_fill
        cell.font = Font(bold=True, size=10)
        cell.border = border
        cell.alignment = Alignment(horizontal="right" if col_idx > 3 else "left", vertical="center")
    ws1.row_dimensions[total_row].height = 18
    _autowidth(ws1, SUMMARY_HEADERS, total_row)
    ws1.freeze_panes = "A2"

    # ── Hoja 2: Details ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Details")
    _write_header(ws2, DETAIL_HEADERS, header_fill, HEADER_FG, border)

    detalle = _detalle_fichajes_rango(db, fecha_von, fecha_bis, empleado_ids)
    for row_idx, r in enumerate(detalle, start=2):
        data = [
            r["personalnr"], r["name"],
            r["datum"].strftime("%d.%m.%Y"),
            r["beginn"].strftime("%H:%M"),
            r["ende"].strftime("%H:%M") if r["ende"] else "",
            r["pause"], r["stunden"], r["kostenstelle"],
        ]
        row_fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if col_idx in (1, 6, 7) else "left", vertical="center")
            cell.font = Font(size=10)
            if row_fill:
                cell.fill = row_fill
    last_row = len(detalle) + 1
    _autowidth(ws2, DETAIL_HEADERS, last_row)
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def preview_reporte_rango(
    db: Session,
    fecha_von: date,
    fecha_bis: date,
    empleado_ids: Optional[list] = None,
) -> dict:
    """Preview JSON del reporte por rango (resumen por empleado + conteos)."""
    resumen, festivos = _resumen_empleados_rango(db, fecha_von, fecha_bis, empleado_ids)
    detalle_count = len(_detalle_fichajes_rango(db, fecha_von, fecha_bis, empleado_ids))

    totals = {
        "planstunden": round(sum(r["planstunden"] for r in resumen), 2),
        "iststunden": round(sum(r["iststunden"] for r in resumen), 2),
        "differenz": round(sum(r["differenz"] for r in resumen), 2),
        "saldo_kumuliert": round(sum(r["saldo_kumuliert"] for r in resumen), 2),
        "krankheitstage": sum(r["krankheitstage"] for r in resumen),
        "urlaubstage": sum(r["urlaubstage"] for r in resumen),
    }
    return {
        "fecha_von": fecha_von.isoformat(),
        "fecha_bis": fecha_bis.isoformat(),
        "total_empleados": len(resumen),
        "total_fichajes": detalle_count,
        "festivos": festivos,
        "empleados": resumen,
        "totales": totals,
    }
