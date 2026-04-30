"""
Import masivo de empleados via Excel/CSV — HG-Plan E
Backup a Telegram — HG-Plan H
"""
import io
import csv
import json
import httpx
import subprocess
from datetime import datetime, date
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session

from ..database import get_db
from ..models.empleado import Empleado, Grupo, CentroCoste, Zeitgruppe
from ..models.audit import AuditLog
from ..services.audit_service import log_action

router = APIRouter(tags=["Import / Export"])

TELEGRAM_BOT_TOKEN = "8631473423:AAFFeU426CXzYdDDQEdvmVhHL48U8cmLGTU"
# Bot sends to itself (saved messages) — admin must /start the bot first
# We'll use a configurable chat_id
TELEGRAM_BACKUP_CHAT_ID = None  # Set via POST /backup/config


# ── Import Excel/CSV ────────────────────────────────────

IMPORT_FIELD_MAP = {
    # Excel column → (model field, type)
    "Systemnummer": ("id_nummer", int),
    "Vorname": ("nombre", str),
    "Nachname": ("apellido", str),
    "Personalnummer": ("personalnummer", int),
    "Benutzer-ID": ("benutzer_id", int),
    "Transponder-ID": ("nfc_tag", str),
    "Beginn der Berechnung": ("beginn_berechnung", "date"),
    "Mandat": ("mandat", str),
    "Firmenbereich": ("firmenbereich", str),
}

BENUTZERSTATUS_MAP = {
    "Admin": 1,
    "Schichtführer": 2,
    "Stv. Schichtführer": 3,
    "Benutzer": 4,
}


@router.post("/import/mitarbeiter")
async def import_mitarbeiter(
    file: UploadFile = File(...),
    dry_run: bool = Query(False, description="Nur validieren, nicht speichern"),
    db: Session = Depends(get_db),
):
    """
    Importiert Mitarbeiter aus Excel (.xlsx) oder CSV (.csv).
    Erwartet Spalten: Systemnummer, Vorname, Nachname, Personalnummer,
    Benutzer-ID, Benutzerstatus, Transponder-ID, Beginn der Berechnung,
    Zeitgruppe, Abteilung, Kostenstelle, Mandat, Firmenbereich
    """
    filename = file.filename or ""
    content = await file.read()

    if filename.endswith(".xlsx"):
        rows = _parse_xlsx(content)
    elif filename.endswith(".csv"):
        rows = _parse_csv(content)
    else:
        raise HTTPException(400, "Nur .xlsx oder .csv Dateien erlaubt")

    if not rows:
        raise HTTPException(400, "Keine Daten in der Datei gefunden")

    # Pre-load lookups
    grupos = {g.nombre: g for g in db.query(Grupo).all()}
    kostenstellen = {k.nombre: k for k in db.query(CentroCoste).all()}
    zeitgruppen = {z.nombre: z for z in db.query(Zeitgruppe).all()}
    existing_ids = {e.id_nummer for e in db.query(Empleado.id_nummer).all()}

    results = {"erstellt": 0, "aktualisiert": 0, "fehler": [], "gesamt": len(rows)}

    for i, row in enumerate(rows, start=2):
        try:
            sys_nr = int(row.get("Systemnummer", 0))
            if not sys_nr:
                results["fehler"].append({"zeile": i, "fehler": "Systemnummer fehlt"})
                continue

            vorname = str(row.get("Vorname", "")).strip()
            nachname = str(row.get("Nachname", "")).strip()
            if not vorname:
                results["fehler"].append({"zeile": i, "fehler": "Vorname fehlt"})
                continue

            # Resolve Abteilung
            abt_name = str(row.get("Abteilung", "")).strip()
            grupo = None
            if abt_name and abt_name != "<Keine>":
                if abt_name not in grupos:
                    g = Grupo(nombre=abt_name)
                    if not dry_run:
                        db.add(g)
                        db.flush()
                    grupos[abt_name] = g
                grupo = grupos[abt_name]

            # Resolve Kostenstelle
            ks_name = str(row.get("Kostenstelle", "")).strip()
            ks = None
            if ks_name and ks_name != "<Keine>":
                if ks_name not in kostenstellen:
                    code = ks_name[:20].upper().replace(" ", "_")
                    c = CentroCoste(codigo=code, nombre=ks_name)
                    if not dry_run:
                        db.add(c)
                        db.flush()
                    kostenstellen[ks_name] = c
                ks = kostenstellen[ks_name]

            # Resolve Zeitgruppe
            zg_name = str(row.get("Zeitgruppe", "")).strip()
            zg = None
            if zg_name and zg_name != "<Keine>":
                if zg_name not in zeitgruppen:
                    tipo = "GLEITZEIT"
                    if "schicht" in zg_name.lower():
                        tipo = "SCHICHT"
                    elif "verwaltung" in zg_name.lower():
                        tipo = "VERWALTUNG"
                    z = Zeitgruppe(nombre=zg_name, tipo=tipo)
                    if not dry_run:
                        db.add(z)
                        db.flush()
                    zeitgruppen[zg_name] = z
                zg = zeitgruppen[zg_name]

            # Parse dates
            beginn = _parse_date(row.get("Beginn der Berechnung"))

            # Build employee data
            emp_data = {
                "id_nummer": sys_nr,
                "nombre": vorname,
                "apellido": nachname,
                "nfc_tag": str(row.get("Transponder-ID", "")).strip() or None,
                "beginn_berechnung": beginn,
                "mandat": str(row.get("Mandat", "<Keine>")).strip(),
                "firmenbereich": str(row.get("Firmenbereich", "<Keine>")).strip(),
            }

            # Personalnummer / Benutzer-ID
            pnr = row.get("Personalnummer")
            if pnr:
                emp_data["personalnummer"] = int(pnr)
            bid = row.get("Benutzer-ID")
            if bid:
                emp_data["benutzer_id"] = int(bid)

            if sys_nr in existing_ids:
                # Update existing
                if not dry_run:
                    emp = db.query(Empleado).filter(Empleado.id_nummer == sys_nr).first()
                    for k, v in emp_data.items():
                        if k != "id_nummer":
                            setattr(emp, k, v)
                    if grupo:
                        emp.grupo_id = grupo.id
                    if ks:
                        emp.kostenstelle_id = ks.id
                    if zg:
                        emp.zeitgruppe_id = zg.id
                results["aktualisiert"] += 1
            else:
                # Create new
                if not dry_run:
                    emp = Empleado(**emp_data, activo=True, fecha_alta=beginn)
                    if grupo:
                        emp.grupo_id = grupo.id
                    if ks:
                        emp.kostenstelle_id = ks.id
                    if zg:
                        emp.zeitgruppe_id = zg.id
                    db.add(emp)
                results["erstellt"] += 1
                existing_ids.add(sys_nr)

        except Exception as ex:
            results["fehler"].append({"zeile": i, "fehler": str(ex)})

    if not dry_run:
        log_action(db, "IMPORT", "empleado",
                   descripcion=f"Import: {results['erstellt']} erstellt, {results['aktualisiert']} aktualisiert",
                   usuario_nick="admin")
        db.commit()

    results["dry_run"] = dry_run
    return results


# ── Backup ───────────────────────────────────────────────

@router.post("/backup/telegram")
def backup_to_telegram(
    chat_id: int = Query(..., description="Telegram Chat-ID für Backup"),
    db: Session = Depends(get_db),
):
    """
    Erstellt ein Datenbank-Backup und sendet es an Telegram.
    """
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"hagemann_backup_{timestamp}.sql"
    filepath = f"/tmp/{filename}"

    # pg_dump
    try:
        result = subprocess.run(
            ["pg_dump", "-h", "postgres", "-U", "postgres", "-d", "neofreight",
             "--schema=hagemann", "-f", filepath],
            env={"PGPASSWORD": "localdev"},
            capture_output=True, text=True, timeout=60,
        )
        if result.returncode != 0:
            raise HTTPException(500, f"pg_dump Fehler: {result.stderr}")
    except FileNotFoundError:
        # pg_dump not available in container — use Python export
        return _python_backup(db, chat_id, timestamp)

    # Send to Telegram
    _send_telegram_file(filepath, filename, chat_id)

    log_action(db, "BACKUP", "system",
               descripcion=f"Backup gesendet an Telegram Chat {chat_id}",
               usuario_nick="system")
    db.commit()

    return {"status": "ok", "datei": filename, "chat_id": chat_id}


def _python_backup(db: Session, chat_id: int, timestamp: str):
    """Fallback: export tables as JSON."""
    from ..models.empleado import Empleado, Grupo, CentroCoste, Zeitgruppe
    from ..models.fichaje import Fichaje
    from ..models.turno import ModeloTurno, PlanTurno

    tables = {
        "empleados": db.query(Empleado).all(),
        "grupos": db.query(Grupo).all(),
        "centros_coste": db.query(CentroCoste).all(),
        "zeitgruppen": db.query(Zeitgruppe).all(),
    }

    data = {}
    for name, rows in tables.items():
        data[name] = [
            {c.name: str(getattr(r, c.name, None)) for c in r.__table__.columns}
            for r in rows
        ]

    filename = f"hagemann_backup_{timestamp}.json"
    filepath = f"/tmp/{filename}"
    with open(filepath, "w") as f:
        json.dump(data, f, indent=2, default=str)

    _send_telegram_file(filepath, filename, chat_id)

    log_action(db, "BACKUP", "system",
               descripcion=f"JSON-Backup gesendet an Telegram Chat {chat_id}",
               usuario_nick="system")
    db.commit()

    return {"status": "ok", "datei": filename, "format": "json", "chat_id": chat_id}


def _send_telegram_file(filepath: str, filename: str, chat_id: int):
    """Send file to Telegram bot."""
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
    with open(filepath, "rb") as f:
        resp = httpx.post(
            url,
            data={"chat_id": chat_id, "caption": f"🗄 Hagemann Backup\n{filename}"},
            files={"document": (filename, f)},
            timeout=30,
        )
    if resp.status_code != 200:
        raise HTTPException(500, f"Telegram Fehler: {resp.text}")


# ── Helpers ──────────────────────────────────────────────

def _parse_xlsx(content: bytes) -> list[dict]:
    """Parse Excel file into list of dicts. Prefers 'Importvorlage' sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb["Importvorlage"] if "Importvorlage" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        if not any(row):
            continue
        result.append({headers[i]: row[i] for i in range(len(headers)) if i < len(row)})
    return result


def _parse_csv(content: bytes) -> list[dict]:
    """Parse CSV file into list of dicts."""
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text), delimiter=";")
    return list(reader)


def _parse_date(val) -> Optional[date]:
    """Parse various date formats."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None
